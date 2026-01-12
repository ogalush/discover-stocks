import streamlit as st
import pandas as pd
import yfinance as yf
from datetime import datetime, timedelta
from utils.common import get_ticker, get_stock_name
from utils.db import get_connection
from io import BytesIO
import mplfinance as mpf
import matplotlib
import os
import zipfile
from matplotlib.font_manager import FontProperties
from functools import lru_cache
from openpyxl.styles import numbers

def init_session_state():
    """セッション状態の初期化"""
    if 'stock_data' not in st.session_state:
        st.session_state['stock_data'] = {}
    if 'charts' not in st.session_state:
        st.session_state['charts'] = {}
    if 'vote_input_codes' not in st.session_state:
        st.session_state['vote_input_codes'] = ""
    if 'direct_input_codes' not in st.session_state:
        st.session_state['direct_input_codes'] = ""
    if 'vote_input_codes_area' not in st.session_state:
        st.session_state['vote_input_codes_area'] = ""
    if 'direct_input_codes_area' not in st.session_state:
        st.session_state['direct_input_codes_area'] = ""

@lru_cache(maxsize=400)
def get_stock_data(stock_code, start_date, end_date):
    """
    株価データを取得する関数（キャッシュ付き）
    
    Parameters:
    stock_code (str): 銘柄コード
    start_date (str): 開始日（YYYY-MM-DD形式）
    end_date (str): 終了日（YYYY-MM-DD形式）
    
    Returns:
    DataFrame: 株価データ
    """
    try:
        # yfinanceのTicker形式に変換
        ticker = get_ticker(stock_code)
        
        # 終了日を翌日にずらす（yfinanceは[start, end)の半開区間）
        end_date_plus_one = (pd.Timestamp(end_date) + pd.Timedelta(days=1)).strftime("%Y-%m-%d")
        
        # 株価データを取得
        df = yf.download(
            ticker,
            start=start_date,
            end=end_date_plus_one,
            progress=False,
            threads=False,
            auto_adjust=True
        )
        
        return df
        
    except Exception as e:
        st.error(f"データ取得中にエラーが発生しました: {str(e)}")
        return pd.DataFrame()

def create_candlestick_chart(df):
    """
    ローソク足チャートを作成する関数
    データ量が250ポイント以上の場合は自動的に折れ線グラフに切り替え
    
    Parameters:
    df (DataFrame): 株価データ
    
    Returns:
    bytes: チャート画像のバイナリデータ
    """
    # インデックスがDateTimeIndexであることを確認
    if not isinstance(df.index, pd.DatetimeIndex):
        df.index = pd.DatetimeIndex(df.index)
        
    # マルチインデックスの場合はレベル0を選択
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
        
    # mplfinanceで必要なOHLCV形式に変換
    df_plot = df[['Open', 'High', 'Low', 'Close', 'Volume']].copy()
    
    # データポイント数に基づいてチャートタイプを決定
    if len(df_plot) > 250:  # 約1年分の取引日
        chart_type = 'line'
    else:
        chart_type = 'candle'
    
    # スタイルの設定
    mpf_style = mpf.make_mpf_style(base_mpf_style='yahoo', marketcolors=mpf.make_marketcolors(
        up='red',
        down='blue',
        edge='inherit',
        wick='inherit',
        volume='inherit',
    ))
    
    # Bytesオブジェクトを作成
    buf = BytesIO()
    
    # チャートを作成
    fig, axes = mpf.plot(
        df_plot,
        type=chart_type,  # 動的にチャートタイプを設定
        volume=True,
        style=mpf_style,
        returnfig=True,
        figsize=(12, 8),
        panel_ratios=(4, 1)
    )
    
    # チャートをバイト形式で保存
    fig.savefig(buf, format='png', bbox_inches='tight')
    buf.seek(0)
    
    return buf, chart_type  # チャートタイプも返す

def get_vote_results_top_n(vote_date, top_n=20):
    """指定日の投票結果上位N件を取得"""
    conn = get_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT stock_code, COUNT(*) as vote_count
            FROM vote
            WHERE vote_date = ?
            GROUP BY stock_code
            ORDER BY vote_count DESC
            LIMIT ?
        """, (vote_date, top_n))
        return cursor.fetchall()  # [(銘柄コード, 投票数), ...]
    finally:
        conn.close()

def show(selected_date):
    st.title("特定銘柄分析ページ")
    
    # セッション状態の初期化
    init_session_state()

    # 最大登録数
    MAX_STOCKS = st.number_input("同時登録最大数", min_value=1, max_value=150, value=50, step=1)

    # 銘柄指定方法の選択
    input_method = st.radio(
        "銘柄指定方法",
        ["投票結果から銘柄を挿入", "銘柄コードをカンマ区切りで入力"],
        index=0,
        horizontal=True,
        key="input_method_radio"
    )

    stock_codes_for_analysis = ""

    if input_method == "投票結果から銘柄を挿入":
        st.markdown("### 投票結果設定")
        col_vote1, col_vote2 = st.columns(2)
        with col_vote1:
            vote_date = st.date_input(
                "投票日",
                value=selected_date,
                min_value=datetime(2020, 1, 1).date(),
                max_value=datetime.now().date(),
                key="vote_date_for_insert"
            )
        with col_vote2:
            insert_count = st.number_input(
                "挿入件数",
                min_value=1,
                max_value=150,
                value=20,
                step=1,
                key="insert_count"
            )
        
        insert_mode = st.radio(
            "挿入方法",
            ["置換（既存をクリア）", "追加（末尾に追加）"],
            horizontal=True,
            key="insert_mode"
        )
        
        if st.button("投票結果を挿入", key="insert_vote_results"):
            vote_date_str = vote_date.strftime("%Y-%m-%d")
            
            try:
                vote_results = get_vote_results_top_n(vote_date_str, insert_count)
                
                if vote_results:
                    # 銘柄コードのみを抽出
                    new_codes = [code for code, _ in vote_results]
                    
                    if insert_mode == "置換（既存をクリア）":
                        new_value = ", ".join(new_codes)
                    else:  # 追加
                        existing_codes = [code.strip() for code in st.session_state['vote_input_codes'].split(",") if code.strip()]
                        # 重複を除いて追加
                        for code in new_codes:
                            if code not in existing_codes:
                                existing_codes.append(code)
                        new_value = ", ".join(existing_codes)
                    
                    # セッション状態を即時更新
                    st.session_state['vote_input_codes'] = new_value
                    st.session_state['vote_input_codes_area'] = new_value
                    st.rerun()
                else:
                    st.warning("指定された日付に投票結果がありません。")
            except Exception as e:
                st.error(f"エラーが発生しました: {str(e)}")

        # 編集可能なテキストエリア
        stock_codes_for_analysis = st.text_area(
            "挿入結果プレビュー（編集可能）",
            height=120,
            key="vote_input_codes_area",
            help=f"最大{MAX_STOCKS}個まで"
        )
        # 入力値をセッション状態に同期
        st.session_state['vote_input_codes'] = stock_codes_for_analysis

    else:
        # 直接入力モード
        stock_codes_for_analysis = st.text_area(
            "銘柄コードをカンマ区切りで入力（例: 7203, 6758）",
            height=120,
            key="direct_input_codes_area",
            help=f"最大{MAX_STOCKS}個まで"
        )
        # 入力値をセッション状態に同期
        st.session_state['direct_input_codes'] = stock_codes_for_analysis

    # 入力された銘柄コードをリスト化（共通処理）
    stock_code_list = [code.strip() for code in stock_codes_for_analysis.split(",") if code.strip()][:MAX_STOCKS]

    # 期間設定モードの選択
    date_mode = st.radio(
        "期間設定モード",
        ["共通設定", "銘柄ごと設定"],
        index=0,  # デフォルトは共通設定
        horizontal=True,
        key="date_mode"
    )

    # 共通の期間設定（共通設定モードまたはデフォルト値として使用）
    col1, col2 = st.columns(2)
    with col1:
        common_start_date = st.date_input(
            "分析開始日（共通）" if date_mode == "共通設定" else "デフォルト開始日",
            value=datetime.now().date() - timedelta(days=120),  # 4ヶ月前
            min_value=datetime(2010, 1, 1).date(),
            max_value=datetime.now().date(),
            key="common_start_date"
        )
    with col2:
        common_end_date = st.date_input(
            "分析終了日（共通）" if date_mode == "共通設定" else "デフォルト終了日",
            value=datetime.now().date(),
            min_value=common_start_date,
            max_value=datetime.now().date(),
            key="common_end_date"
        )

    # 銘柄ごとの期間設定（銘柄ごと設定モードの場合）
    stock_dates = {}
    if date_mode == "銘柄ごと設定" and stock_code_list:
        st.write("**銘柄ごとの期間設定**")
        for code in stock_code_list:
            stock_name = get_stock_name(code)
            with st.expander(f"{stock_name}({code}) の期間設定", expanded=False):
                col_s, col_e = st.columns(2)
                with col_s:
                    start_date = st.date_input(
                        f"開始日",
                        value=common_start_date,
                        min_value=datetime(2010, 1, 1).date(),
                        max_value=datetime.now().date(),
                        key=f"start_date_{code}"
                    )
                with col_e:
                    end_date = st.date_input(
                        f"終了日",
                        value=common_end_date,
                        min_value=start_date,
                        max_value=datetime.now().date(),
                        key=f"end_date_{code}"
                    )
                stock_dates[code] = (start_date, end_date)

    # データ取得・表示
    if st.button("データ取得"):
        progress_bar = st.progress(0)
        total_stocks = len(stock_code_list)

        # 新しいデータ取得時にはセッション状態をリセット
        st.session_state['stock_data'] = {}
        st.session_state['charts'] = {}
        
        for i, code in enumerate(stock_code_list):
            try:
                # 進捗バーの更新
                progress = (i + 1) / total_stocks
                progress_bar.progress(progress)

                # 期間設定モードに応じて開始日・終了日を決定
                if date_mode == "銘柄ごと設定" and code in stock_dates:
                    start_date, end_date = stock_dates[code]
                else:
                    start_date, end_date = common_start_date, common_end_date

                start_date_str = start_date.strftime("%Y-%m-%d")
                end_date_str = end_date.strftime("%Y-%m-%d")
                
                # キャッシュ付きのデータ取得関数を使用
                df = get_stock_data(code, start_date_str, end_date_str)
                
                if not df.empty:
                    # セッション状態にデータを保存
                    st.session_state['stock_data'][code] = df
                    
                    # ローソク足チャートの作成と保存
                    buf, chart_type = create_candlestick_chart(df)
                    st.session_state['charts'][code] = {
                        'buf': buf,
                        'type': chart_type
                    }
                else:
                    st.warning(f"{code} のデータが取得できませんでした。")
                
            except Exception as e:
                st.error(f"{code} のデータ取得中にエラーが発生しました: {str(e)}")
                continue
        
        progress_bar.progress(1.0)
    
    # 保存されたデータを表示
    if st.session_state['stock_data']:
        # 一括ダウンロードボタンを追加
        if len(st.session_state['stock_data']) > 0:
            # Excelファイルの作成
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # 各銘柄のデータをシートとして追加
                for code, df in st.session_state['stock_data'].items():
                    sheet_name = f"{code}_{get_stock_name(code)}"
                    # シート名が長すぎる場合は短縮
                    if len(sheet_name) > 31:  # Excelのシート名の最大長
                        sheet_name = f"{code}_{get_stock_name(code)[:20]}"
                    
                    # データフレームをExcelに書き込み
                    df.to_excel(writer, sheet_name=sheet_name, index=True)
                    
                    # ワークシートの取得
                    worksheet = writer.sheets[sheet_name]
                    
                    # 列幅の自動調整
                    for idx, col in enumerate(df.columns):
                        max_length = max(
                            df[col].astype(str).apply(len).max(),
                            len(str(col))
                        )
                        worksheet.column_dimensions[chr(65 + idx + 1)].width = max_length + 2  # +1 for index column
                    
                    # インデックス列（日付）の幅も調整
                    max_date_length = max(
                        max(len(str(date)) for date in df.index),
                        len('Date')
                    )
                    worksheet.column_dimensions['A'].width = max_date_length + 2
                    
                    # 日付列の書式を設定
                    for row in range(2, len(df) + 2):  # 2から始まる（ヘッダー行の後）
                        cell = worksheet.cell(row=row, column=1)  # A列
                        cell.number_format = 'yyyy/m/d'
                    
                    # TradingViewのURLを追加
                    url = f'https://jp.tradingview.com/chart/?symbol={code}'
                    worksheet.cell(row=1, column=len(df.columns) + 2).value = 'TradingView URL'
                    worksheet.cell(row=1, column=len(df.columns) + 2).hyperlink = url
                    worksheet.cell(row=1, column=len(df.columns) + 2).style = 'Hyperlink'
            
            # Excelファイルのダウンロードボタン
            st.download_button(
                label="一括ダウンロード（Excel）",
                data=output.getvalue(),
                file_name=f"stock_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_excel"
            )
            
            # CSV一括ダウンロード（ZIPファイル）
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for code, df in st.session_state['stock_data'].items():
                    csv_data = df.to_csv().encode('utf-8-sig')
                    # UTF-8フラグを設定してファイル名の文字化けを防止
                    file_name = f"{code}_{get_stock_name(code)}_stock_data.csv"
                    zip_info = zipfile.ZipInfo(file_name)
                    zip_info.flag_bits |= 0x800  # UTF-8フラグ（bit 11）を設定
                    zip_info.compress_type = zipfile.ZIP_DEFLATED
                    zip_file.writestr(zip_info, csv_data)
            
            st.download_button(
                label="一括ダウンロード（CSV/ZIP）",
                data=zip_buffer.getvalue(),
                file_name=f"stock_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                mime="application/zip",
                key="download_csv_zip"
            )
        
        # 個別のデータ表示
        for code, df in st.session_state['stock_data'].items():
            st.subheader(f"{get_stock_name(code)} ({code})")
            st.write("【株価データ】")
            st.dataframe(df)
            st.write("【チャート】")
            
            # データフレームの構造を確認して適切に処理
            if isinstance(df.columns, pd.MultiIndex):
                # マルチインデックスの場合はレベル0を選択
                chart_df = df.copy()
                chart_df.columns = chart_df.columns.get_level_values(0)  # 最初のレベルのカラム名を取得
                st.line_chart(chart_df[["Open", "Close", "High", "Low"]])
            else:
                # 通常のカラム構造の場合
                st.line_chart(df[["Open", "Close", "High", "Low"]])
            
            # ローソク足チャートの表示
            if code in st.session_state['charts']:
                chart_info = st.session_state['charts'][code]
                chart_type_text = "ローソク足" if chart_info['type'] == 'candle' else "折れ線"
                st.image(chart_info['buf'], caption=f"{get_stock_name(code)} ({code}) - {chart_type_text}チャート", use_container_width=True)
            
            # CSVダウンロード
            csv = df.to_csv().encode('utf-8-sig')
            st.download_button(
                label="CSVダウンロード",
                data=csv,
                file_name=f"{code}_stock_data.csv",
                mime="text/csv",
                key=f"download_{code}"  # 一意のキーを設定
            )